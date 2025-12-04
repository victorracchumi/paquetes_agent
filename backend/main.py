# Backend API for Recepción de Paquetes - v2.0
import os
from typing import Optional, Literal
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field
from datetime import datetime
from openpyxl import load_workbook
import requests
import msal

from dotenv import load_dotenv
from pathlib import Path

# Cargar .env desde la raíz del proyecto (un nivel arriba de backend/)
env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

EXCEL_PATH = os.getenv("EXCEL_PATH", "./data/recepcion_paquetes.xlsx")

TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
GRAPH_SENDER_UPN = os.getenv("GRAPH_SENDER_UPN")

TEAMS_WEBHOOK_URL = os.getenv("TEAMS_WEBHOOK_URL", "")  # optional channel notification

app = FastAPI(title="Recepción de Paquetes — Backend", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class PackageIn(BaseModel):
    sucursal: str
    recepcionista: str
    proveedor: str
    tipoDocumento: str
    numeroDocumento: str
    destinatarioNombre: str
    destinatarioEmail: EmailStr
    medioNotificacion: Literal["Correo", "Teams", "Ambos"]
    observaciones: Optional[str] = ""
    adjuntoUrl: Optional[str] = ""
    fechaRecepcion: str = Field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))
    horaRecepcion: str = Field(default_factory=lambda: datetime.now().strftime("%H:%M:%S"))
    codigoRetiro: str
    # Campos específicos para cheques
    montoCheque: Optional[str] = ""
    fechaVencimientoCheque: Optional[str] = ""

class PackageOut(BaseModel):
    id: str
    estado: str

def ensure_excel_headers(path: str):
    # The file should exist (created earlier). If not, create it with headers.
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RecepcionLog"
        headers = [
            "FechaRecepcion","HoraRecepcion","Sucursal","Recepcionista",
            "Proveedor","TipoDocumento","NumeroDocumento",
            "DestinatarioNombre","DestinatarioEmail","MedioNotificacion",
            "CodigoRetiro","Estado","FechaNotificacion","DestinatarioConfirmo",
            "FechaRetiro","EntregadoA","Observaciones","AdjuntoUrl"
        ]
        ws.append(headers)
        wb.save(path)

def append_row_to_excel(path: str, values: list):
    ensure_excel_headers(path)
    wb = load_workbook(path)
    ws = wb["RecepcionLog"]
    ws.append(values)
    wb.save(path)

def msal_acquire_token() -> Optional[str]:
    if not (TENANT_ID and CLIENT_ID and CLIENT_SECRET):
        return None
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app_msal = msal.ConfidentialClientApplication(
        CLIENT_ID, authority=authority, client_credential=CLIENT_SECRET
    )
    result = app_msal.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    return result.get("access_token")

def send_email_graph(to_email: str, subject: str, html_body: str) -> bool:
    token = msal_acquire_token()
    if not token:
        return False
    url = f"https://graph.microsoft.com/v1.0/users/{GRAPH_SENDER_UPN}/sendMail"
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": html_body},
            "toRecipients": [{"emailAddress": {"address": to_email}}],
        },
        "saveToSentItems": "true",
    }
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    return r.status_code in (202, 200)

def notify_teams_webhook(text: str) -> bool:
    if not TEAMS_WEBHOOK_URL:
        return False
    payload = {"text": text}
    try:
        r = requests.post(TEAMS_WEBHOOK_URL, json=payload, timeout=15)
        return r.status_code in (200, 204)
    except Exception:
        return False

def format_email_html(pkg: PackageIn) -> str:
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.4;
                color: #333;
                max-width: 550px;
                margin: 0 auto;
            }}
            .header {{
                background: #4a5568;
                color: white;
                padding: 15px;
                text-align: center;
            }}
            .header h2 {{
                margin: 0;
                font-size: 18px;
            }}
            .content {{
                padding: 20px;
                background: #f9f9f9;
            }}
            .info-table {{
                width: 100%;
                margin: 15px 0;
                background: white;
                border: 1px solid #ddd;
            }}
            .info-table td {{
                padding: 8px 12px;
                border-bottom: 1px solid #eee;
                font-size: 14px;
            }}
            .info-table tr:last-child td {{
                border-bottom: none;
            }}
            .info-table td:first-child {{
                font-weight: 600;
                color: #4a5568;
                width: 35%;
            }}
            .important {{
                background: #fff3cd;
                border-left: 3px solid #ff9800;
                padding: 12px;
                margin: 15px 0;
                font-size: 13px;
            }}
            .important strong {{
                color: #e65100;
            }}
            .footer {{
                margin-top: 15px;
                font-size: 13px;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2>📦 Recepción de Correspondencia</h2>
        </div>

        <div class="content">
            <p>Estimado(a) <strong>{pkg.destinatarioNombre}</strong>,</p>
            <p>Ha llegado correspondencia a su nombre:</p>

            <table class="info-table">
                <tr>
                    <td>🚚 Proveedor</td>
                    <td>{pkg.proveedor}</td>
                </tr>
                <tr>
                    <td>📄 Tipo de Documento</td>
                    <td>{pkg.tipoDocumento}</td>
                </tr>
                <tr>
                    <td>🔢 Número de Documento</td>
                    <td>{pkg.numeroDocumento}</td>
                </tr>
                {"" if pkg.tipoDocumento != "Cheque" or not pkg.montoCheque else f'''
                <tr>
                    <td>💰 Monto del Cheque</td>
                    <td>{pkg.montoCheque}</td>
                </tr>
                '''}
                {"" if pkg.tipoDocumento != "Cheque" or not pkg.fechaVencimientoCheque else f'''
                <tr>
                    <td>📆 Fecha de Vencimiento</td>
                    <td>{pkg.fechaVencimientoCheque}</td>
                </tr>
                '''}
                <tr>
                    <td>📅 Fecha de Recepción</td>
                    <td>{pkg.fechaRecepcion}</td>
                </tr>
                <tr>
                    <td>🕐 Hora de Recepción</td>
                    <td>{pkg.horaRecepcion}</td>
                </tr>
            </table>

            <div class="important">
                <strong>🚩 Importante:</strong> Retiro mismo día de notificación<br>
                • Lunes a Jueves: hasta 18:00 hrs<br>
                • Viernes: hasta 17:00 hrs
            </div>

            <div class="footer">
                <p>De antemano, muchas gracias.</p>
                <p><strong>Rudth Nuñez</strong><br>
                Recepcionista - Multiaceros S.A.<br>
                📧 recepcion@multiaceros.cl</p>
            </div>
        </div>
    </body>
    </html>
    """

@app.get("/health")
def health():
    return {"ok": True}

@app.post("/register", response_model=PackageOut)
def register_package(pkg: PackageIn):
    excel_row = [
        pkg.fechaRecepcion, pkg.horaRecepcion, pkg.sucursal, pkg.recepcionista,
        pkg.proveedor, pkg.tipoDocumento, pkg.numeroDocumento,
        pkg.destinatarioNombre, pkg.destinatarioEmail, pkg.medioNotificacion,
        pkg.codigoRetiro, "Pendiente", "", False,
        "", "", pkg.observaciones or "", pkg.adjuntoUrl or ""
    ]
    try:
        append_row_to_excel(EXCEL_PATH, excel_row)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error registrando en Excel: {e}")

    estado = "Pendiente"
    notified = False

    if pkg.medioNotificacion in ("Correo", "Ambos"):
        ok_mail = send_email_graph(
            pkg.destinatarioEmail,
            f"Recepción de paquete — {pkg.destinatarioNombre} — {pkg.sucursal}",
            format_email_html(pkg),
        )
        notified = notified or ok_mail

    if pkg.medioNotificacion in ("Teams", "Ambos"):
        text = f"📦 Paquete para {pkg.destinatarioNombre} ({pkg.destinatarioEmail}). Proveedor: {pkg.proveedor}. Doc: {pkg.tipoDocumento} {pkg.numeroDocumento}. Código: {pkg.codigoRetiro}. Sucursal: {pkg.sucursal}."
        ok_teams = notify_teams_webhook(text)
        notified = notified or ok_teams

    if notified:
        estado = "Notificado"

    id_simple = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{pkg.codigoRetiro}"
    return {"id": id_simple, "estado": estado}

class ReminderRequest(BaseModel):
    email: EmailStr
    nombre: str

def format_reminder_email_html(nombre: str) -> str:
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.4;
                color: #333;
                max-width: 550px;
                margin: 0 auto;
            }}
            .header {{
                background: #e65100;
                color: white;
                padding: 15px;
                text-align: center;
            }}
            .header h2 {{
                margin: 0;
                font-size: 18px;
            }}
            .content {{
                padding: 20px;
                background: #fff8e1;
            }}
            .reminder {{
                background: #ffccbc;
                border-left: 4px solid #d84315;
                padding: 15px;
                margin: 15px 0;
                font-size: 14px;
            }}
            .reminder strong {{
                color: #bf360c;
                font-size: 16px;
            }}
            .footer {{
                margin-top: 15px;
                font-size: 13px;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2>⏰ Recordatorio de Retiro</h2>
        </div>

        <div class="content">
            <p>Estimado(a) <strong>{nombre}</strong>,</p>

            <div class="reminder">
                <strong>🚨 Recordatorio:</strong><br><br>
                Tiene correspondencia pendiente de retiro en Recepción.<br><br>
                Por favor, recuerde retirarla durante el horario de atención:<br>
                • Lunes a Jueves: hasta 18:00 hrs<br>
                • Viernes: hasta 17:00 hrs
            </div>

            <div class="footer">
                <p>De antemano, muchas gracias.</p>
                <p><strong>Rudth Nuñez</strong><br>
                Recepcionista - Multiaceros S.A.<br>
                📧 recepcion@multiaceros.cl</p>
            </div>
        </div>
    </body>
    </html>
    """

@app.get("/packages")
def get_packages():
    """
    Endpoint para obtener todos los paquetes registrados del Excel.
    """
    try:
        if not os.path.exists(EXCEL_PATH):
            return {"packages": []}

        wb = load_workbook(EXCEL_PATH)
        ws = wb["RecepcionLog"]

        packages = []
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Si tiene fecha de recepción
                package = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        package[header] = row[i]
                packages.append(package)

        return {"packages": packages}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error leyendo paquetes: {e}")

@app.post("/send-reminder")
def send_reminder(req: ReminderRequest):
    """
    Endpoint para enviar recordatorio de retiro a un usuario.
    Puede ser llamado por el chatbot cuando el usuario lo solicita.
    """
    try:
        ok_mail = send_email_graph(
            req.email,
            f"⏰ Recordatorio: Correspondencia pendiente - {req.nombre}",
            format_reminder_email_html(req.nombre),
        )

        if ok_mail:
            return {"success": True, "message": f"Recordatorio enviado a {req.email}"}
        else:
            return {"success": False, "message": "No se pudo enviar el recordatorio"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error enviando recordatorio: {e}")

@app.get("/search-users")
def search_users(query: str):
    """
    Busca usuarios y grupos de distribución en Azure AD / Microsoft 365 usando Microsoft Graph API.
    Retorna una lista de usuarios/grupos con displayName y email.

    Args:
        query: Texto de búsqueda (nombre, apellido, email parcial, nombre de grupo)

    Returns:
        Lista de usuarios/grupos: [{"displayName": "Juan Pérez", "email": "jperez@empresa.cl", "type": "user"},
                                    {"displayName": "Cobranzas", "email": "cobranzas@empresa.cl", "type": "group"}, ...]
    """
    if not query or len(query) < 2:
        return {"users": []}

    try:
        token = msal_acquire_token()
        if not token:
            raise HTTPException(status_code=500, detail="No se pudo obtener token de Microsoft Graph")

        # Buscar usuarios con Graph API
        # Obtener TODOS los usuarios y filtrar localmente (más flexible que startswith)
        url = "https://graph.microsoft.com/v1.0/users"
        headers = {
            "Authorization": f"Bearer {token}",
            "ConsistencyLevel": "eventual"
        }

        # Obtener todos los usuarios (o usar $top para limitar)
        params = {
            "$select": "displayName,mail,userPrincipalName",
            "$top": "999"  # Obtener hasta 999 usuarios
        }

        # Buscar usuarios
        response_users = requests.get(url, headers=headers, params=params, timeout=10)

        results = []
        query_lower = query.lower()

        # Procesar usuarios con filtrado local
        if response_users.status_code == 200:
            data = response_users.json()

            for user in data.get("value", []):
                display_name = user.get("displayName", "")
                email = user.get("mail") or user.get("userPrincipalName", "")

                # Filtrar localmente: buscar query en displayName o email (case-insensitive, contains)
                if email and display_name:
                    if (query_lower in display_name.lower() or
                        query_lower in email.lower()):
                        results.append({
                            "displayName": display_name,
                            "email": email,
                            "type": "user"
                        })

        # Buscar grupos de distribución
        # Obtenemos TODOS los grupos con email y filtramos localmente
        groups_url = "https://graph.microsoft.com/v1.0/groups"
        groups_params = {
            "$filter": "mailEnabled eq true",
            "$select": "displayName,mail,mailEnabled,securityEnabled",
            "$top": "999"
        }

        try:
            response_groups = requests.get(groups_url, headers=headers, params=groups_params, timeout=10)

            # Procesar grupos
            if response_groups.status_code == 200:
                groups_data = response_groups.json()
                query_lower = query.lower()

                for group in groups_data.get("value", []):
                    display_name = group.get("displayName", "")
                    email = group.get("mail", "")
                    mail_enabled = group.get("mailEnabled", False)

                    # Filtrar localmente: buscar query en el nombre o email (case-insensitive y con contains)
                    if email and display_name and mail_enabled:
                        if query_lower in display_name.lower() or query_lower in email.lower():
                            results.append({
                                "displayName": f"📧 {display_name} (Grupo)",
                                "email": email,
                                "type": "group"
                            })
            else:
                # Si falla la búsqueda de grupos, registrar error pero continuar
                print(f"Error al buscar grupos: {response_groups.status_code} - {response_groups.text}")
        except Exception as e:
            # Si hay error en búsqueda de grupos, solo registrar pero no fallar
            print(f"Excepción al buscar grupos: {str(e)}")

        # Limitar resultados a máximo 15 (10 usuarios + 5 grupos máximo)
        if results:
            # Separar usuarios y grupos
            usuarios = [r for r in results if r.get("type") == "user"]
            grupos = [r for r in results if r.get("type") == "group"]

            # Limitar a 10 usuarios y 5 grupos
            resultados_finales = usuarios[:10] + grupos[:5]

            return {"users": resultados_finales}
        else:
            # Si $filter falla, intentar con $search (requiere permisos adicionales)
            params_search = {
                "$search": f'"displayName:{query}" OR "mail:{query}"',
                "$select": "displayName,mail,userPrincipalName",
                "$top": "10"
            }

            response_search = requests.get(url, headers=headers, params=params_search, timeout=10)

            if response_search.status_code == 200:
                data = response_search.json()
                users = []

                for user in data.get("value", []):
                    display_name = user.get("displayName", "")
                    email = user.get("mail") or user.get("userPrincipalName", "")

                    if email and display_name:
                        users.append({
                            "displayName": display_name,
                            "email": email
                        })

                return {"users": users}
            else:
                raise HTTPException(
                    status_code=response.status_code,
                    detail=f"Error de Microsoft Graph: {response.text}"
                )

    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR EN SEARCH-USERS: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error buscando usuarios: {str(e)}")

@app.get("/debug-groups")
def debug_groups():
    """
    Endpoint de depuración para listar TODOS los grupos con email.
    Útil para ver qué grupos están disponibles.
    """
    try:
        token = msal_acquire_token()
        if not token:
            raise HTTPException(status_code=500, detail="No se pudo obtener token")

        url = "https://graph.microsoft.com/v1.0/groups"
        headers = {
            "Authorization": f"Bearer {token}",
            "ConsistencyLevel": "eventual"
        }
        params = {
            "$select": "displayName,mail,mailEnabled,securityEnabled,groupTypes",
            "$top": "50",
            "$filter": "mailEnabled eq true"
        }

        response = requests.get(url, headers=headers, params=params, timeout=10)

        if response.status_code == 200:
            data = response.json()
            groups = []

            for group in data.get("value", []):
                groups.append({
                    "displayName": group.get("displayName", ""),
                    "mail": group.get("mail", ""),
                    "mailEnabled": group.get("mailEnabled", False),
                    "securityEnabled": group.get("securityEnabled", False),
                    "groupTypes": group.get("groupTypes", [])
                })

            return {"total": len(groups), "groups": groups}
        else:
            return {"error": f"Status {response.status_code}", "detail": response.text}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {e}")
